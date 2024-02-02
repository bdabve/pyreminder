#!/usr/bin/env python3
# -*- coding: utf-8 -*-
#
# author        : el3arbi el3arbi_@email.com
# created       :
#
# description   :
# usage         :
# ----------------------------------------------------------------------------
import sqlite3
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo


def make_query(db_name, query, params=()):
    conn = sqlite3.connect(db_name)
    curs = conn.cursor()
    try:
        curs.execute(query, params)
    except sqlite3.Error as err:
        print(err)
    else:
        desc = [desc[0] for desc in curs.description]
        rows = curs.fetchall()
        return (desc, rows)


def xlsx_table(desc, rows, filename):
    wb = Workbook()
    ws = wb.active
    ws.append(desc)
    for row in rows:
        ws.append(row)

    # exe of ref : "A1:E5"
    # ws.min_row, ws.max_row        : minimum and maximum row containing data
    # ws.min_column, ws.max_column  : minimum and maximum column containing data
    from_cell = ws.cell(ws.min_row, ws.min_column).coordinate
    to_cell = ws.cell(ws.max_row, ws.max_column).coordinate
    cell_range = '{}:{}'.format(from_cell, to_cell)                 # create range for row that contain data
    table = Table(displayName="Table1", ref=cell_range)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False,
                           showRowStripes=True, showColumnStripes=True)
    table.tableStyleInfo = style
    ws.add_table(table)
    wb.save(filename)
    print('(+) done writint to: {}'.format(filename))


if __name__ == '__main__':
    db_name = '/home/dabve/app_django/inv/inv/db.sqlite3'
    query = 'SELECT * FROM magasin_article WHERE code LIKE ? LIMIT ?'
    params = ['%BHS%', 10]
    desc, rows = make_query(db_name, query, params)
    xlsx_table(desc, rows, 'BHS_ARTICLE.xlsx')
